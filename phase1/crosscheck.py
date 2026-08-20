"""
Phase 1 · Join-key coverage and cross-source agreement.

The build plan asserts CMR reference is the natural primary key across PDFs,
storefront CSVs, crosstab workbooks and pack shots. This module tests that
assertion, and — more usefully — tests whether the sources *agree* where they
overlap.

Cross-source disagreement is the thing that quietly destroys trust in a
research tool. Better to find it now, in a report, than in a client meeting.

Output: key_coverage.json, key_coverage.csv, disagreements.csv
"""
from __future__ import annotations

import csv
import json
import re
import warnings
from collections import defaultdict
from pathlib import Path

import openpyxl

import config

warnings.filterwarnings("ignore")

CMR_IN_NAME = re.compile(r"(\d{6})")


def _read_csv_any(path: Path) -> list[dict]:
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with path.open(encoding=enc, newline="") as fh:
                return list(csv.DictReader(fh))
        except UnicodeDecodeError:
            continue
    return []


def collect_pdf_keys() -> dict[str, dict]:
    recs_path = config.OUT_DIR / "pdf_records.json"
    if not recs_path.exists():
        return {}
    out = {}
    for r in json.loads(recs_path.read_text(encoding="utf-8")):
        if r["cmr_ref"]:
            out[r["cmr_ref"]] = {
                "set": r["session_set"],
                "product_name": r["product_name"],
                "manufacturer": r["manufacturer"],
                "price": r["price_gbp"],
                "size": r["size"],
                "score": r["score_out_of_50"],
                "category": r["category"],
            }
    return out


def collect_storefront_keys() -> tuple[dict, dict]:
    """WIX export (scores) and ECWID export (category/brand/storage)."""
    wix, ecwid = {}, {}
    for p in config.FFX_REPORTS.rglob("*.csv"):
        if any(x in config.SKIP_DIR_NAMES for x in p.parts):
            continue
        rows = _read_csv_any(p)
        if not rows:
            continue
        cols = {c.lower().strip(): c for c in rows[0].keys() if c}
        if "cmr" in cols:
            for r in rows:
                key = str(r.get(cols["cmr"], "")).strip()
                if key:
                    wix[key] = {
                        "name": (r.get(cols.get("name", ""), "") or "").strip(),
                        "score": (r.get(cols.get("score", ""), "") or "").strip(),
                        "rating": (r.get(cols.get("rating", ""), "") or "").strip(),
                        "price": (r.get(cols.get("price", ""), "") or "").strip(),
                        "size": (r.get(cols.get("size", ""), "") or "").strip(),
                        "src": p.name,
                    }
        elif "sku" in cols:
            for r in rows:
                key = str(r.get(cols["sku"], "")).strip()
                if key:
                    ecwid[key] = {
                        "name": (r.get(cols.get("name", ""), "") or "").strip(),
                        "category": (r.get(cols.get("category 1", ""), "") or "").strip(),
                        "brand": (r.get(cols.get("brand", ""), "") or "").strip(),
                        "storage": (r.get(cols.get("upc", ""), "") or "").strip(),
                        "src": p.name,
                    }
    return wix, ecwid


def collect_photo_keys() -> dict[str, list[str]]:
    photos = defaultdict(list)
    for p in config.FFX_REPORTS.rglob("*"):
        if p.suffix.lower() not in (".jpg", ".jpeg", ".png"):
            continue
        if any(x in config.SKIP_DIR_NAMES for x in p.parts):
            continue
        m = CMR_IN_NAME.search(p.stem)
        if m:
            photos[m.group(1)].append(p.name)
        else:
            photos["__unmatched__"].append(p.name)
    return dict(photos)


def collect_crosstab_keys() -> dict[str, list[str]]:
    """CMR refs referenced by the session workbooks' OT Sheet (test order map)."""
    found = defaultdict(list)
    for p in config.FFX_REPORTS.rglob("*.xls*"):
        if any(x in config.SKIP_DIR_NAMES for x in p.parts):
            continue
        try:
            wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
        except Exception:                                     # noqa: BLE001
            continue
        try:
            if "OT Sheet" not in wb.sheetnames:
                continue
            for row in wb["OT Sheet"].iter_rows(values_only=True):
                for v in row:
                    if v is None:
                        continue
                    m = re.fullmatch(r"(\d{6})", str(v).strip())
                    if m:
                        found[m.group(1)].append(p.name)
        finally:
            wb.close()
    return dict(found)


def _norm_money(v) -> float | None:
    if v in (None, ""):
        return None
    t = re.sub(r"[^\d.]", "", str(v))
    try:
        return round(float(t), 2)
    except ValueError:
        return None


def run() -> dict:
    out = config.ensure_out()
    print("  cross-checking join keys across PDFs, CSVs, workbooks and photos...")

    pdfs = collect_pdf_keys()
    wix, ecwid = collect_storefront_keys()
    photos = collect_photo_keys()
    crosstabs = collect_crosstab_keys()

    unmatched_photos = photos.pop("__unmatched__", [])
    universe = set(pdfs) | set(wix) | set(ecwid) | set(photos) | set(crosstabs)

    rows = []
    for k in sorted(universe):
        rows.append({
            "cmr_ref": k,
            "in_pdf": k in pdfs,
            "in_wix": k in wix,
            "in_ecwid": k in ecwid,
            "has_photo": k in photos,
            "in_ot_sheet": k in crosstabs,
            "n_sources": sum([k in pdfs, k in wix, k in ecwid,
                              k in photos, k in crosstabs]),
            "product": (pdfs.get(k, {}).get("product_name")
                        or wix.get(k, {}).get("name")
                        or ecwid.get(k, {}).get("name") or ""),
        })

    # --- agreement between PDF and WIX where both exist -------------------
    disagreements = []
    both = set(pdfs) & set(wix)
    for k in sorted(both):
        p, w = pdfs[k], wix[k]
        try:
            w_score = int(float(w["score"])) if w["score"] else None
        except ValueError:
            w_score = None
        if w_score is not None and p["score"] is not None and w_score != p["score"]:
            disagreements.append({"cmr_ref": k, "field": "score",
                                  "pdf": p["score"], "wix": w_score,
                                  "product": p["product_name"]})
        pp, wp = _norm_money(p["price"]), _norm_money(w["price"])
        if pp is not None and wp is not None and abs(pp - wp) > 0.001:
            disagreements.append({"cmr_ref": k, "field": "price",
                                  "pdf": pp, "wix": wp,
                                  "product": p["product_name"]})
        ps = (p["size"] or "").lower().replace(" ", "")
        ws = (w["size"] or "").lower().replace(" ", "")
        if ps and ws and ps != ws:
            disagreements.append({"cmr_ref": k, "field": "size",
                                  "pdf": p["size"], "wix": w["size"],
                                  "product": p["product_name"]})

    summary = {
        "distinct_cmr_refs_total": len(universe),
        "coverage": {
            "pdf_reports": len(pdfs),
            "wix_export": len(wix),
            "ecwid_export": len(ecwid),
            "pack_photos": len(photos),
            "ot_sheet": len(crosstabs),
        },
        "in_all_five_sources": sum(1 for r in rows if r["n_sources"] == 5),
        "in_one_source_only": sum(1 for r in rows if r["n_sources"] == 1),
        "pdf_and_wix_overlap": len(both),
        "photos_unmatched_to_a_cmr_ref": unmatched_photos,
        "agreement_checked_pairs": len(both),
        "disagreements": len(disagreements),
        "disagreement_by_field": {
            f: sum(1 for d in disagreements if d["field"] == f)
            for f in ("score", "price", "size")},
        "orphans": {
            "pdf_without_wix": sorted(set(pdfs) - set(wix)),
            "wix_without_pdf": sorted(set(wix) - set(pdfs)),
            "photo_without_pdf": sorted(set(photos) - set(pdfs)),
        },
    }

    with (out / "key_coverage.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)

    with (out / "disagreements.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["cmr_ref", "field", "pdf", "wix", "product"])
        w.writeheader()
        w.writerows(disagreements)

    (out / "key_coverage.json").write_text(
        json.dumps(summary, indent=2), encoding="utf-8")
    return summary


if __name__ == "__main__":
    print(json.dumps(run(), indent=2))
