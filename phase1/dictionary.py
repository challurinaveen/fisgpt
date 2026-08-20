"""
Phase 1 · Measure dictionary.

`Foodfax Database details.xlsx` is 37 rows and is the semantic layer for the
whole FFX database — it says what each question is, which products it is asked
of, which statistic is reported, which report templates consume it and which
clients commissioned it. The build plan calls it the highest value-per-byte
file in the corpus.

This module loads it and does the part that makes it usable by a text-to-SQL
prompt: maps each business-language measure onto the *physical columns* that
actually carry it in `Norm Data` and `FFX DATA`. Anything that fails to map is
reported rather than dropped — an unmapped measure is a question the tool will
silently be unable to answer.

Output: measure_dictionary.json, measure_mapping.csv
"""
from __future__ import annotations

import csv
import json
import re
import warnings
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl

import config

warnings.filterwarnings("ignore")

# Business vocabulary that differs between the dictionary and the column
# headers. These are the aliases the physical schema actually uses.
ALIASES = {
    "aroma": ["smell"],
    "smell": ["aroma"],
    "texture": ["texture/mouthfeel", "mouthfeel"],
    "overall quality": ["overall impression / quality", "overall impression", "quality"],
    "value for what you pay": ["value for money"],
    "market comparison": ["better than"],
    "initial appeal": ["appeal", "want to try"],
    "likelihood to notice": ["notice"],
    "exciting new idea": ["new&different", "new & different"],
    "pre-test interest in purchase": ["pre pi", "pre-test pi", "pre-test interest in purchasing"],
    "full price post test pi": ["priced pi", "full price post test pi"],
    "brand": ["brand fit"],
    "recommend to a friend": ["recommend to a friend", "product promoter score"],
}

STAT_PATTERNS = [
    (re.compile(r"top\s*2\s*box", re.I), "T2B"),
    (re.compile(r"top\s*3\s*box", re.I), "T3B"),
    (re.compile(r"top\s*4\s*box", re.I), "T4B"),
    (re.compile(r"bot(?:tom)?\s*2\s*box", re.I), "B2B"),
    (re.compile(r"mean\s*score", re.I), "MEAN"),
    (re.compile(r"all\s*proportions", re.I), "DIST"),
    (re.compile(r"calcul", re.I), "CALC"),
    (re.compile(r"definitely\s*/\s*probably", re.I), "DEFPROB"),
]

QNUM = re.compile(r"^\s*(\d+[a-z]?)\.\s*(.+)$", re.I)


def _norm(s: str) -> str:
    s = re.sub(r"^\s*\d+[a-z]?\.\s*", "", str(s or ""))
    s = s.replace("\n", " ")
    s = re.sub(r"\(.*?\)", " ", s)
    s = re.sub(r"[^a-z0-9 &/]", " ", s.lower())
    return re.sub(r"\s+", " ", s).strip()


def _similar(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()


def load_dictionary() -> list[dict]:
    path = config.FFX_INFO / "Foodfax Database details.xlsx"
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []

    header = [str(c or "").strip() for c in rows[0]]

    def resolve(*fragments) -> int | None:
        """
        Locate a column by meaning rather than exact header text. The sheet's
        own header reads "Additional Projects / Clients"; an exact-match lookup
        for "... / Client" silently returns nothing and the clients column is
        quietly dropped. Match on a distinctive fragment instead.
        """
        for i, h in enumerate(header):
            low = h.lower()
            if any(f in low for f in fragments):
                return i
        return None

    idx = {
        "question": resolve("question"),
        "asked_of": resolve("products asked", "asked for"),
        "comments": resolve("comment"),
        "reports": resolve("reports needed", "report"),
        "clients": resolve("client", "additional project"),
    }
    missing = [k for k, v in idx.items() if v is None]
    if missing:
        print(f"    warning: dictionary columns not found: {missing} "
              f"(header was {header})")

    def cell(r, key, default=""):
        i = idx.get(key)
        return str(r[i]).strip() if i is not None and i < len(r) and r[i] is not None else default

    measures = []
    for r in rows[1:]:
        q = cell(r, "question")
        if not q:
            continue
        m = QNUM.match(q)
        comments = cell(r, "comments")
        stats = [tag for pat, tag in STAT_PATTERNS if pat.search(comments)]
        measures.append({
            "question_code": m.group(1) if m else "",
            "question": (m.group(2) if m else q).replace("\n", " ").strip(),
            "question_raw": q.replace("\n", " ").strip(),
            "asked_of": cell(r, "asked_of"),
            "statistics_raw": comments,
            "statistics": stats or ["UNSPECIFIED"],
            "reports_needed_for": cell(r, "reports"),
            "clients": cell(r, "clients"),
            "is_calculated": "CALC" in stats,
            "normalised": _norm(m.group(2) if m else q),
        })
    return measures


def physical_columns() -> dict[str, list[dict]]:
    """Pull column names from the profiled workbooks (needs profile_excel first)."""
    prof = config.OUT_DIR / "excel_profile.json"
    if not prof.exists():
        return {}
    data = json.loads(prof.read_text(encoding="utf-8"))
    wanted = {
        ("Foodfax Database ARCHIVE.xlsm", "Norm Data"): "norm_data",
        ("FFX DB Update pre 2021.xlsx", "Sheet1"): "ffx_pre_2021",
    }
    cols: dict[str, list[dict]] = {}
    for pr in data.get("profiles", []):
        for s in pr.get("sheets", []):
            key = wanted.get((pr["name"], s["sheet"]))
            if s["sheet"] == "FFX DATA":
                key = "ffx_data"
            if not key:
                continue
            cols.setdefault(key, [])
            for c in s["columns"]:
                if c["name"]:
                    cols[key].append({"name": c["name"], "index": c["index"],
                                      "type": c["inferred_type"]})
    return cols


def map_measures(measures: list[dict], cols: dict[str, list[dict]]) -> list[dict]:
    for msr in measures:
        target = msr["normalised"]
        candidates = [target, *ALIASES.get(target, [])]
        matches = []
        for table, clist in cols.items():
            for c in clist:
                cn = _norm(c["name"])
                if not cn:
                    continue
                # strip the T2B%/ prefix the percentage block uses
                variant = "T2B" if re.match(r"^t2b", cn) else (
                          "T4B" if re.match(r"^t4b", cn) else "MEAN")
                base = re.sub(r"^t[234]b\s*", "", cn)
                best = max(
                    (max(_similar(cand, base),
                         1.0 if cand == base else 0.0) for cand in candidates),
                    default=0.0)
                if best >= 0.86:
                    matches.append({"table": table, "column": c["name"],
                                    "column_index": c["index"],
                                    "variant": variant, "score": round(best, 3)})
        matches.sort(key=lambda m: -m["score"])
        msr["physical_columns"] = matches
        msr["mapped"] = bool(matches)
        msr["variants_available"] = sorted({m["variant"] for m in matches})
        msr["tables"] = sorted({m["table"] for m in matches})
    return measures


def run() -> dict:
    out = config.ensure_out()
    print("  loading measure dictionary and mapping to physical columns...")

    measures = load_dictionary()
    cols = physical_columns()
    measures = map_measures(measures, cols)

    mapped = [m for m in measures if m["mapped"]]
    unmapped = [m for m in measures if not m["mapped"]]

    summary = {
        "measures_total": len(measures),
        "measures_mapped": len(mapped),
        "measures_unmapped": len(unmapped),
        "mapping_rate_pct": round(100 * len(mapped) / max(1, len(measures)), 1),
        "calculated_measures": sum(1 for m in measures if m["is_calculated"]),
        "physical_tables_scanned": {k: len(v) for k, v in cols.items()},
        "unmapped_detail": [
            {"code": m["question_code"], "question": m["question"],
             "reason": "calculated field, no source column" if m["is_calculated"]
                       else "no column within similarity threshold"}
            for m in unmapped],
        "statistics_distribution": {
            s: sum(1 for m in measures if s in m["statistics"])
            for s in ("MEAN", "T2B", "T3B", "T4B", "B2B", "DIST", "CALC",
                      "DEFPROB", "UNSPECIFIED")},
        "measures_with_named_clients": sum(1 for m in measures if m["clients"]),
        "measures_tied_to_report_templates": sum(
            1 for m in measures if m["reports_needed_for"]),
    }

    (out / "measure_dictionary.json").write_text(
        json.dumps({"summary": summary, "measures": measures}, indent=2),
        encoding="utf-8")

    with (out / "measure_mapping.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["question_code", "question", "asked_of", "statistics",
                    "reports_needed_for", "clients", "mapped", "tables",
                    "variants", "top_column_match"])
        for m in measures:
            w.writerow([m["question_code"], m["question"], m["asked_of"],
                        "|".join(m["statistics"]), m["reports_needed_for"],
                        m["clients"], m["mapped"], "|".join(m["tables"]),
                        "|".join(m["variants_available"]),
                        m["physical_columns"][0]["column"] if m["physical_columns"] else ""])

    return summary


if __name__ == "__main__":
    print(json.dumps(run(), indent=2))
