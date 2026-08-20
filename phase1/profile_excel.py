"""
Phase 1 · Workbook profiling with shape detection.

The point of this module is Finding C's corollary: a generic "first row with
three strings is the header" heuristic silently corrupts `Norm Data`, whose
real header is row 7 behind six rows of summary statistics. So we do not
guess once — we score every candidate row, classify the sheet's shape, and
then check ourselves against the known-shape registry in config.py.

Where the blind detection and the registry disagree, that is a finding, not a
crash. It gets reported.

Output: excel_profile.json, excel_columns.csv, excel_shapes.csv
"""
from __future__ import annotations

import csv
import json
import re
import warnings
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import openpyxl

import config

warnings.filterwarnings("ignore")

SCAN_ROWS = 20          # how deep to look for a header
SAMPLE_ROWS = 400       # rows sampled per sheet for column typing
MAX_COLS = 200

STATS_LABELS = {
    "mean scores", "mean score", "minimum", "maximum",
    "standard deviation", "standard error of mean", "std dev",
    "number of products", "count", "average",
}
BANNER_HINT = re.compile(r"\bn\s*=\s*\d+", re.I)


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------
def _s(v) -> str:
    """Cell to string, with embedded newlines collapsed.

    Header cells in `Norm Data` are authored multi-line ("T2B%
Notice"), which
    breaks CSV round-tripping and SQL identifier generation downstream."""
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _nonempty(row) -> list:
    return [v for v in row if v is not None and _s(v) != ""]


def _type_of(v) -> str:
    if v is None or _s(v) == "":
        return "null"
    if isinstance(v, bool):
        return "bool"
    if isinstance(v, (int, float)):
        return "number"
    if isinstance(v, (datetime, date)):
        return "date"
    t = _s(v)
    if re.fullmatch(r"-?\d{1,3}(,\d{3})*(\.\d+)?%?", t):
        return "number"
    return "text"


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------
def score_header_row(rows: list[tuple], i: int, ncols: int) -> float:
    """
    Score row `i` as a header candidate. Higher is better.

    A header row is: mostly text, mostly filled, mostly distinct, and — the
    signal that actually separates row 7 from row 1 in Norm Data — followed by
    rows that are mostly *numeric*.
    """
    row = rows[i]
    cells = _nonempty(row)
    if len(cells) < 3:
        return -99.0

    texts = [c for c in cells if isinstance(c, str) and _s(c)]
    str_ratio = len(texts) / len(cells)
    fill_ratio = len(cells) / max(1, ncols)
    distinct = len({_s(c).lower() for c in cells}) / len(cells)

    below = rows[i + 1: i + 8]
    if below:
        numeric_frac = []
        for r in below:
            vals = _nonempty(r)
            if not vals:
                continue
            numeric_frac.append(
                sum(1 for v in vals if _type_of(v) == "number") / len(vals))
        below_numeric = sum(numeric_frac) / len(numeric_frac) if numeric_frac else 0.0
        below_fill = sum(len(_nonempty(r)) for r in below) / max(1, len(below) * ncols)
    else:
        below_numeric = below_fill = 0.0

    score = (str_ratio * 3.0) + (fill_ratio * 2.5) + (distinct * 2.0) \
            + (below_numeric * 3.0) + (below_fill * 1.5)

    # A row carrying summary-statistic labels is a stats row, not a header.
    joined = " ".join(_s(c).lower() for c in texts)
    if any(lbl in joined for lbl in STATS_LABELS):
        score -= 6.0
    # Rows the sheet author used as a title usually have one or two cells only.
    if len(cells) <= 2:
        score -= 4.0
    # Earlier rows win ties, so a genuine row-1 header is not beaten by row 2.
    score -= i * 0.06
    return score


def detect_header(rows: list[tuple], ncols: int) -> tuple[int, float, list[float]]:
    """Returns (1-based header row index, confidence 0..1, all scores)."""
    scores = [score_header_row(rows, i, ncols) for i in range(min(SCAN_ROWS, len(rows)))]
    if not scores or max(scores) < 0:
        return 0, 0.0, scores
    best = max(range(len(scores)), key=lambda i: scores[i])
    ordered = sorted((s for s in scores if s > -99), reverse=True)
    margin = (ordered[0] - ordered[1]) if len(ordered) > 1 else ordered[0]
    confidence = max(0.0, min(1.0, margin / 3.0))
    return best + 1, round(confidence, 3), scores


# ---------------------------------------------------------------------------
# Shape classification
# ---------------------------------------------------------------------------
def classify_shape(rows: list[tuple], header_row: int, ncols: int,
                   sheet: str) -> tuple[str, int]:
    """
    Returns (shape, header_row). The header row can be corrected here: a
    two-row header is detected as a *pair*, and the pair's upper row is the
    real header even when the blind scorer preferred the lower one.

      flat             — header on one row, data below
      stats_block      — summary statistics sit above the real header
      two_row_header   — block label on row N, attribute on row N+1, and the
                         attribute names REPEAT across blocks, so each column
                         needs its block prefix to stay unique
      coalesce_header  — the two rows are complementary: whichever of them
                         names a given column is that column's name
      banner_crosstab  — header cells carry base sizes ("n = 45")
      report_surface   — a rendering template, not data
      key_value        — small config/lookup sheet
      empty
    """
    if not rows or all(not _nonempty(r) for r in rows):
        return "empty", header_row

    if sheet.startswith(config.REPORT_SURFACE_PREFIXES):
        return "report_surface", header_row

    head_blob = " ".join(
        _s(v) for r in rows[:min(10, len(rows))] for v in r if v is not None)
    if BANNER_HINT.search(head_blob):
        return "banner_crosstab", header_row

    if header_row > 1:
        above = " ".join(
            _s(v).lower() for r in rows[:header_row - 1] for v in r if v is not None)
        if any(lbl in above for lbl in STATS_LABELS):
            return "stats_block", header_row

    # Consider the detected row and the one above it as a possible pair.
    for base in sorted({max(1, header_row - 1), header_row}):
        if base < 1 or base >= len(rows):
            continue
        upper, lower = rows[base - 1], rows[base]
        low_cells = _nonempty(lower)
        up_cells = _nonempty(upper)
        if len(low_cells) < 3 or not up_cells:
            continue

        low_text = [c for c in low_cells if isinstance(c, str) and _s(c)]
        if len(low_text) / len(low_cells) < 0.6:
            continue                      # lower row is data, not names

        # Does the lower row repeat itself? If the same attribute name appears
        # under more than one block, the block prefix is load-bearing.
        seen = [_s(c).lower() for c in low_text]
        lower_has_duplicates = len(seen) != len(set(seen))

        # Where the lower row names a column, is the cell above it a usable
        # name, a numeric code, or blank?
        above_text = above_numeric = above_blank = 0
        for ci in range(ncols):
            if ci >= len(lower) or not _s(lower[ci]):
                continue
            a = upper[ci] if ci < len(upper) else None
            if a is None or _s(a) == "":
                above_blank += 1
            elif _type_of(a) == "number":
                above_numeric += 1
            else:
                above_text += 1
        named = above_text + above_numeric + above_blank
        if named < 3:
            continue

        upper_unusable = (above_numeric + above_blank) / named

        if lower_has_duplicates and above_text >= 1:
            # Repeating attributes under sparse block labels — forward-fill
            # the block across and prefix (FFX DB pre-2021: Appearance appears
            # once under "MS" and again under "Percentage (T2B ...)").
            return "two_row_header", base
        if upper_unusable > 0.75 and not lower_has_duplicates:
            # The row above carries codes or nothing where the row below
            # carries the real names (Retailer 5 Year Summary).
            return "coalesce_header", base

    if ncols <= 3 and len(rows) <= 60:
        return "key_value", header_row
    return "flat", header_row


# ---------------------------------------------------------------------------
# Column profiling
# ---------------------------------------------------------------------------
def profile_columns(rows, header_row, ncols, shape) -> list[dict]:
    if header_row < 1 or header_row > len(rows):
        return []

    names = list(rows[header_row - 1])[:ncols]

    # Complementary headers: take whichever of the two rows names the column.
    if shape == "coalesce_header" and header_row < len(rows):
        sub = list(rows[header_row])[:ncols]
        names = [
            (_s(b) or _s(a)) for a, b in zip(
                list(names) + [None] * ncols, list(sub) + [None] * ncols)
        ][:ncols]
        data_start = header_row + 1
        body = rows[data_start: data_start + SAMPLE_ROWS]
        out = []
        for ci in range(ncols):
            col = [r[ci] if ci < len(r) else None for r in body]
            types = Counter(_type_of(v) for v in col)
            n = len(col) or 1
            vals = [v for v in col if _type_of(v) != "null"]
            dominant = types.most_common(1)[0][0] if types else "null"
            if dominant == "null" and len(types) > 1:
                dominant = types.most_common(2)[1][0]
            out.append({
                "index": ci + 1,
                "name": names[ci] if ci < len(names) else "",
                "inferred_type": dominant,
                "null_pct": round(100 * types.get("null", 0) / n, 1),
                "distinct_sampled": len({_s(v) for v in vals}),
                "samples": [_s(v)[:40] for v in vals[:3]],
            })
        return out

    # For a two-row header, forward-fill the block label across and pair it
    # with the attribute name beneath (the FFX DB pre-2021 quirk, §04).
    if shape == "two_row_header" and header_row < len(rows):
        sub = list(rows[header_row])[:ncols]
        filled, last = [], ""
        for v in names:
            t = _s(v)
            if t:
                last = t
            filled.append(last)
        names = [
            f"{b} · {_s(s)}" if _s(s) and b else (_s(s) or b)
            for b, s in zip(filled, sub)
        ]
        data_start = header_row + 1
    else:
        names = [_s(n) for n in names]
        data_start = header_row

    body = rows[data_start: data_start + SAMPLE_ROWS]
    out = []
    for ci in range(ncols):
        col = [r[ci] if ci < len(r) else None for r in body]
        types = Counter(_type_of(v) for v in col)
        n = len(col) or 1
        nulls = types.get("null", 0)
        vals = [v for v in col if _type_of(v) != "null"]
        dominant = types.most_common(1)[0][0] if types else "null"
        if dominant == "null" and len(types) > 1:
            dominant = types.most_common(2)[1][0]
        samples = [_s(v)[:40] for v in vals[:3]]
        out.append({
            "index": ci + 1,
            "name": names[ci] if ci < len(names) else "",
            "inferred_type": dominant,
            "null_pct": round(100 * nulls / n, 1),
            "distinct_sampled": len({_s(v) for v in vals}),
            "samples": samples,
        })
    return out


# ---------------------------------------------------------------------------
# Sheet + workbook
# ---------------------------------------------------------------------------
def profile_sheet(ws, filename: str) -> dict:
    rows: list[tuple] = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        rows.append(r)
        if i >= max(SCAN_ROWS, SAMPLE_ROWS) + 25:
            break

    ncols = min(ws.max_column or 0, MAX_COLS)
    total_rows = ws.max_row or 0

    header_row, confidence, _ = detect_header(rows, ncols)
    shape, header_row = classify_shape(rows, header_row, ncols, ws.title)

    if shape == "report_surface":
        cols = []
    else:
        cols = profile_columns(rows, header_row, ncols, shape)

    known = config.KNOWN_SHAPES.get((filename, ws.title))
    check = "not_registered"
    if known:
        ok_shape = known["shape"] == shape
        ok_row = known["header_row"] == header_row
        check = "match" if (ok_shape and ok_row) else (
            f"MISMATCH expected shape={known['shape']} row={known['header_row']}, "
            f"got shape={shape} row={header_row}")

    return {
        "sheet": ws.title,
        "rows": total_rows,
        "cols": ncols,
        "shape": shape,
        "header_row": header_row,
        "header_confidence": confidence,
        "registry_check": check,
        "ingest": (
            shape not in ("report_surface", "empty")
            and (ws.title in config.SESSION_WORKBOOK_WHITELIST
                 or not ws.title.startswith(config.REPORT_SURFACE_PREFIXES))
        ),
        "columns": cols,
    }


def profile_workbook(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheets = [profile_sheet(wb[sn], path.name) for sn in wb.sheetnames]
    finally:
        wb.close()
    return {
        "file": str(path.relative_to(config.CORPUS_ROOT)).replace("\\", "/"),
        "name": path.name,
        "size_bytes": path.stat().st_size,
        "n_sheets": len(sheets),
        "sheets": sheets,
    }


def target_workbooks() -> list[Path]:
    """Live-zone workbooks only — the Archive copy is a proven duplicate."""
    found = []
    for root in (config.LIVE_DATA, config.PROJECT_DOCS):
        if not root.exists():
            continue
        for p in sorted(root.rglob("*")):
            if p.suffix.lower() in (".xlsx", ".xlsm") and \
               not any(part in config.SKIP_DIR_NAMES for part in p.parts):
                found.append(p)
    return found


def run() -> dict:
    out = config.ensure_out()
    books = target_workbooks()
    print(f"  profiling {len(books)} workbooks...")

    profiles, failures = [], []
    for p in books:
        try:
            profiles.append(profile_workbook(p))
            print(f"    ok  {p.name}")
        except Exception as e:                                # noqa: BLE001
            failures.append({"file": p.name, "error": f"{type(e).__name__}: {e}"})
            print(f"    ERR {p.name}: {type(e).__name__}")

    (out / "excel_profile.json").write_text(
        json.dumps({"profiles": profiles, "failures": failures}, indent=2, default=str),
        encoding="utf-8")

    with (out / "excel_shapes.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["file", "sheet", "rows", "cols", "shape", "header_row",
                    "confidence", "ingest", "registry_check"])
        for pr in profiles:
            for s in pr["sheets"]:
                w.writerow([pr["name"], s["sheet"], s["rows"], s["cols"],
                            s["shape"], s["header_row"], s["header_confidence"],
                            s["ingest"], s["registry_check"]])

    with (out / "excel_columns.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["file", "sheet", "col_index", "column_name", "type",
                    "null_pct", "distinct_sampled", "sample_values"])
        for pr in profiles:
            for s in pr["sheets"]:
                for c in s["columns"]:
                    w.writerow([pr["name"], s["sheet"], c["index"], c["name"],
                                c["inferred_type"], c["null_pct"],
                                c["distinct_sampled"], " | ".join(c["samples"])])

    all_sheets = [s for pr in profiles for s in pr["sheets"]]
    mismatches = [
        {"file": pr["name"], "sheet": s["sheet"], "detail": s["registry_check"]}
        for pr in profiles for s in pr["sheets"]
        if s["registry_check"].startswith("MISMATCH")
    ]
    low_conf = [
        {"file": pr["name"], "sheet": s["sheet"],
         "header_row": s["header_row"], "confidence": s["header_confidence"]}
        for pr in profiles for s in pr["sheets"]
        if s["ingest"] and s["header_confidence"] < 0.30
    ]

    summary = {
        "workbooks_profiled": len(profiles),
        "workbooks_failed": len(failures),
        "sheets_total": len(all_sheets),
        "sheets_to_ingest": sum(1 for s in all_sheets if s["ingest"]),
        "sheets_skipped_report_surface": sum(
            1 for s in all_sheets if s["shape"] == "report_surface"),
        "shapes": dict(Counter(s["shape"] for s in all_sheets).most_common()),
        "registry_mismatches": mismatches,
        "low_confidence_headers": low_conf,
        "failures": failures,
    }
    (out / "excel_profile_summary.json").write_text(
        json.dumps(summary, indent=2), encoding="utf-8")
    return summary


if __name__ == "__main__":
    print(json.dumps(run(), indent=2))
