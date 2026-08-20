"""
Phase 2 · Parse Historic Products (13,278 × 22, flat header row 1).

The pre-2020 product archive. Contains dimensions only — no measure values.
Covers 2002-2019 across FFX, GOLA, NPA, BEA, GAS, OMNI project types and
41 project code names (which map to undisclosed clients — kickoff Q4).

Outputs:
  - product_test DataFrame (13,278 rows)
"""
from __future__ import annotations

import re
import pandas as pd
from openpyxl import load_workbook

import config

ARCHIVE_PATH = config.FFX_EXCEL / "Foodfax Database ARCHIVE.xlsm"


def parse_historic_products() -> pd.DataFrame:
    """Parse Historic Products sheet → product_test DataFrame."""
    print("  loading Historic Products...")
    wb = load_workbook(str(ARCHIVE_PATH), read_only=True, data_only=True)
    ws = wb["Historic Products"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = [_clean(c) for c in rows[0]]
    data = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        d = dict(zip(header, row))
        data.append(d)

    raw_df = pd.DataFrame(data)
    print(f"    {len(raw_df)} rows")

    products_df = pd.DataFrame({
        "unique_reference":   raw_df.get("Unique Reference"),
        "cmr_ref":            raw_df.get("Report", pd.Series(dtype="object")).apply(_str_or_none),
        "prodref":            raw_df.get("PRODREF", pd.Series(dtype="object")).apply(_str_or_none),
        "project_code":       raw_df.get("Project", pd.Series(dtype="object")).apply(_str_or_none),
        "project_name":       raw_df.get("Project Name"),
        "project_type":       raw_df.get("Project Type"),
        "product_name":       raw_df.get("Prodname"),
        "category_code":      raw_df.get("Category code", pd.Series(dtype="object")).apply(_str_or_none),
        "category_name":      raw_df.get("Category"),
        "manufacturer_code":  raw_df.get("Manufacturer Code", pd.Series(dtype="object")).apply(_str_or_none),
        "manufacturer_name":  raw_df.get("MANUFACTURER"),
        "own_label_or_brand": raw_df.get("OWNLABEL"),
        "test_year":          raw_df.get("YEAR", pd.Series(dtype="object")).apply(_int_or_none),
        "test_date":          raw_df.get("DATE").apply(_date_or_none),
        "session_set":        pd.Series([None] * len(raw_df)),
        "test_order":         pd.Series([None] * len(raw_df), dtype="Int64"),
        "price_gbp":          raw_df.get("Price").apply(_float_or_none),
        "pack_size":          raw_df.get("Weight/Size"),
        "storage":            raw_df.get("Storage"),
        "sweet_or_savoury":   raw_df.get("Sweet or Savoury"),
        "fax_type":           raw_df.get("FAXTYPE"),
        "tier":               raw_df.get("Premium Standard Value"),
        "sample_size":        pd.Series([None] * len(raw_df), dtype="Int64"),
        "attribute_scale":    pd.Series([None] * len(raw_df), dtype="Int64"),
        "source_table":       pd.Series(["historic_products"] * len(raw_df)),
        "source_file":        pd.Series(["Foodfax Database ARCHIVE.xlsm"] * len(raw_df)),
    })

    return products_df


def _clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _str_or_none(v):
    if v is None:
        return None
    return str(v).strip() or None


def _float_or_none(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _int_or_none(v):
    if v is None:
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _date_or_none(v):
    if v is None:
        return None
    import datetime
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v).strip()
    if s.lower() in ("n/a", "na", "", "date", "none") or len(s) < 4:
        return None
    return s
