"""
Phase 2 · Parse FFX DB Update pre 2021 (5,811 × 61, two-row header).

The pre-2021 supplement with actual measure values. The two-row header uses
block prefixes: columns 18-33 are "MS · {attribute}" (mean scores) and columns
34-61 are "Percentage (T2B - 5 & 7 PS / T3B 9 PS) · {attribute}" (T2B%).

Phase 1 already handled the header correctly and stored the prefixed column
names in excel_columns.csv. Here we read with openpyxl using those same names.

Outputs:
  - product_test DataFrame (5,811 rows)
  - measure_value DataFrame (melted long form)
"""
from __future__ import annotations

import re
import pandas as pd
from openpyxl import load_workbook

import config

PRE2021_PATH = config.FFX_EXCEL / "FFX DB Update pre 2021.xlsx"

# Dimension columns (cols 1-17)
DIM_COLS = {
    "Project No", "Project Name", "Product / Concept", "Blind/Branded",
    "OwnLabel/Brand", "Priced/Unpriced", "Attribute Scale (5/7/9)",
    "Retail/Food Service", "Tier", "Product Code", "Product Brand",
    "Product Name", "Category No", "Category Name", "Sample Size",
    "Year", "Date",
}


def _parse_variant(col_name: str) -> tuple[str, str] | None:
    """
    Parse block-prefixed column name to (measure_name, variant).

    "MS · Taste" → ("Taste", "MEAN")
    "Percentage (T2B ...) · Taste" → ("Taste", "T2B")
    """
    col = col_name.strip()
    if col.startswith("MS · "):
        return col[5:].strip(), "MEAN"
    if col.startswith("Percentage"):
        # "Percentage (T2B - 5 & 7 PS / T3B 9 PS) · Appearance"
        m = re.search(r"·\s*(.+)$", col)
        if m:
            return m.group(1).strip(), "T2B"
        return None  # bare "Percentage (T2B...)" with no attribute — skip
    return None


def parse_pre2021() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Parse pre-2021 DB → (product_test, measure_value) DataFrames."""
    print("  loading FFX DB Update pre 2021...")
    wb = load_workbook(str(PRE2021_PATH), read_only=True, data_only=True)
    ws = wb["Sheet1"]

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(all_rows) < 3:
        raise ValueError(f"Pre-2021 DB has only {len(all_rows)} rows")

    # Two-row header: row 1 has block labels, row 2 has attribute names
    # Phase 1's profiler built prefixed names. We do the same here.
    row1 = [_clean(c) for c in all_rows[0]]
    row2 = [_clean(c) for c in all_rows[1]]

    # Build composite header with block prefix
    header = []
    current_block = ""
    for i, (top, bottom) in enumerate(zip(row1, row2)):
        if top:
            current_block = top
        if bottom:
            # For dim columns, the block IS the column name
            if current_block == bottom or not current_block:
                header.append(bottom)
            else:
                header.append(f"{current_block} · {bottom}")
        else:
            header.append(current_block if current_block else f"_col_{i}")

    data_rows = all_rows[2:]
    print(f"    {len(data_rows)} data rows, {len(header)} columns")

    # Build raw DataFrame
    products = []
    for row in data_rows:
        if not row or row[0] is None:
            continue
        d = {}
        for col_idx, val in enumerate(row):
            if col_idx < len(header):
                d[header[col_idx]] = val
        products.append(d)

    raw_df = pd.DataFrame(products)
    print(f"    {len(raw_df)} non-empty rows")

    # Product test DataFrame
    products_df = pd.DataFrame({
        "unique_reference":   raw_df.get("Product Code"),
        "cmr_ref":            pd.Series([None] * len(raw_df)),
        "prodref":            pd.Series([None] * len(raw_df)),
        "project_code":       raw_df.get("Project No"),
        "project_name":       raw_df.get("Project Name"),
        "project_type":       pd.Series(["FFX"] * len(raw_df)),  # all FFX in this table
        "product_name":       raw_df.get("Product Name"),
        "category_code":      raw_df.get("Category No", pd.Series(dtype="object")).apply(_str_or_none),
        "category_name":      raw_df.get("Category Name"),
        "manufacturer_code":  pd.Series([None] * len(raw_df)),
        "manufacturer_name":  raw_df.get("Product Brand"),
        "own_label_or_brand": raw_df.get("OwnLabel/Brand"),
        "test_year":          raw_df.get("Year", pd.Series(dtype="object")).apply(_year_or_none),
        "test_date":          raw_df.get("Date").apply(_date_or_none),
        "session_set":        pd.Series([None] * len(raw_df)),
        "test_order":         pd.Series([None] * len(raw_df), dtype="Int64"),
        "price_gbp":          pd.Series([None] * len(raw_df)),  # not in this table
        "pack_size":          pd.Series([None] * len(raw_df)),
        "storage":            pd.Series([None] * len(raw_df)),
        "sweet_or_savoury":   pd.Series([None] * len(raw_df)),
        "fax_type":           pd.Series([None] * len(raw_df)),
        "tier":               raw_df.get("Tier"),
        "sample_size":        raw_df.get("Sample Size", pd.Series(dtype="object")).apply(_int_or_none),
        "attribute_scale":    raw_df.get("Attribute Scale (5/7/9)", pd.Series(dtype="object")).apply(_int_or_none),
        "source_table":       pd.Series(["ffx_pre_2021"] * len(raw_df)),
        "source_file":        pd.Series(["FFX DB Update pre 2021.xlsx"] * len(raw_df)),
    })

    # ── Melt measures ───────────────────────────────────────────────────
    measure_col_indices = []
    for idx, col_name in enumerate(header):
        parsed = _parse_variant(col_name)
        if parsed is not None:
            measure_col_indices.append((idx, col_name, parsed[0], parsed[1]))

    measures = []
    row_idx = 0
    for row in data_rows:
        if not row or row[0] is None:
            continue
        for col_idx, col_name, measure_name, variant in measure_col_indices:
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None:
                continue
            try:
                fval = float(val)
            except (ValueError, TypeError):
                continue
            measures.append({
                "product_row_idx": row_idx,
                "measure_name": measure_name,
                "variant": variant,
                "value": fval,
                "source_table": "ffx_pre_2021",
            })
        row_idx += 1

    measures_df = pd.DataFrame(measures)
    print(f"    {len(measures_df)} measure values melted")

    return products_df, measures_df


def _clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _str_or_none(v):
    if v is None:
        return None
    return str(v).strip() or None


def _int_or_none(v):
    if v is None:
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _year_or_none(v):
    """Extract year from various formats: 2020, datetime(2020,1,1), etc."""
    if v is None:
        return None
    import datetime
    if isinstance(v, datetime.datetime):
        return v.year
    if isinstance(v, datetime.date):
        return v.year
    try:
        n = int(float(str(v)))
        if 1990 <= n <= 2030:
            return n
    except (ValueError, TypeError):
        pass
    return None


def _date_or_none(v):
    if v is None:
        return None
    import datetime
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v).strip()
    if s.lower() in ("n/a", "na", "", "date", "none") or len(s) < 4:
        return None
    return s
