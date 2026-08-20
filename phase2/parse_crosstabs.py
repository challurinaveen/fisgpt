"""
Phase 2 · Parse FFX Tables banner crosstabs → crosstab_cell.

Each FFX Tables workbook has a `Data` sheet (~11,625 rows × 22 cols) containing
every question cut by Gender / Age / Lifestage / SEG / Region / NPS, with base
sizes and significance letters embedded in the header row.

Header format: "Male A n = 11"  →  banner_level="Male", banner_letter="A", base_n=11
Row format: first column is the question/response text; remaining columns are
either percentages (0.0-1.0) or significance letters ("b", "BC").

The banner groups are identified by the column headers:
  Cols 2-3: Gender (NET, Male, Female)
  Cols 4-6: Age (18-34, 35-54, 55+)
  Cols 7-9: Lifestage (Pre Family, Family Combined, Post Family)
  Cols 10-11: SEG (ABC1, C2DE)
  Cols 12-14: Region (North, Midlands, South)
  Cols 15-17: NPS (Promoters, Passive, Detractors)
  Cols 18+: Purchase Frequency (varies)
"""
from __future__ import annotations

import re
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

import config


# Banner group assignment by column position pattern
BANNER_GROUPS = [
    (["NET"],                                    "Total"),
    (["Male", "Female"],                         "Gender"),
    (["18 - 34", "35 - 54", "55 +", "55+"],     "Age Groupings"),
    (["Pre Family", "Family Combined",
      "Post Family"],                            "Lifestage"),
    (["ABC1", "C2DE"],                           "SEG"),
    (["North", "Midlands", "South"],             "Region"),
    (["Promoters", "Passive", "Detractors"],     "NPS"),
    (["Fortnightly", "Monthly", "Weekly",
      "Less Frequent", "More Frequent"],         "Purchase Frequency"),
]

# Regex to parse header like "Male A n = 11" or "NET A n = 45"
RE_HEADER = re.compile(
    r"^(.+?)\s+([A-Z])\s+n\s*=\s*(\d+)$", re.IGNORECASE
)


def _parse_header(h: str) -> dict:
    """Parse 'Male A n = 11' → {level, letter, base_n}."""
    m = RE_HEADER.match(h.strip())
    if not m:
        return {"level": h.strip(), "letter": None, "base_n": None}
    return {
        "level": m.group(1).strip(),
        "letter": m.group(2).upper(),
        "base_n": int(m.group(3)),
    }


def _assign_banner_group(level: str) -> str:
    """Assign a banner group based on the level name."""
    for keywords, group in BANNER_GROUPS:
        for kw in keywords:
            if kw.lower() in level.lower():
                return group
    return "Other"


def parse_crosstab_file(path: Path, session_set: str) -> pd.DataFrame:
    """Parse one FFX Tables workbook → crosstab_cell DataFrame."""
    wb = load_workbook(str(path), read_only=True, data_only=True)

    # Find the Data sheet
    sheet_name = None
    for name in wb.sheetnames:
        if name.lower() == "data":
            sheet_name = name
            break
    if not sheet_name:
        wb.close()
        return pd.DataFrame()

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(all_rows) < 10:
        return pd.DataFrame()

    # Find the header row — look for rows containing "n =" pattern
    header_row_idx = None
    for i, row in enumerate(all_rows[:10]):
        if row and any(
            isinstance(c, str) and "n =" in str(c).lower()
            for c in row if c
        ):
            header_row_idx = i
            break

    if header_row_idx is None:
        print(f"    WARNING: no header found in {path.name}")
        return pd.DataFrame()

    header_raw = all_rows[header_row_idx]

    # Parse banner columns
    banner_cols = []
    for col_idx, h in enumerate(header_raw):
        if col_idx == 0 or h is None:
            continue
        parsed = _parse_header(str(h))
        parsed["col_idx"] = col_idx
        parsed["banner_group"] = _assign_banner_group(parsed["level"])
        banner_cols.append(parsed)

    # Parse data rows
    data_rows = all_rows[header_row_idx + 1:]
    cells = []
    current_question = ""

    for row in data_rows:
        if not row:
            continue

        # First column is question text or response option
        label = str(row[0]).strip() if row[0] is not None else ""
        if not label:
            continue

        # Detect question vs response pattern:
        # Questions tend to be longer; responses are "Top 2 box", "Definitely not", etc.
        # Use heuristic: if most data columns have values, it's a response row
        has_values = sum(
            1 for bc in banner_cols
            if bc["col_idx"] < len(row) and row[bc["col_idx"]] is not None
        )

        if has_values < len(banner_cols) * 0.3:
            # Mostly empty — this is a question header row
            current_question = label
            continue

        response_option = label

        for bc in banner_cols:
            ci = bc["col_idx"]
            if ci >= len(row):
                continue
            val = row[ci]
            if val is None:
                continue

            # Value can be a number (percentage) or a string (sig letter)
            pct = None
            sig = None
            if isinstance(val, (int, float)):
                pct = float(val)
            elif isinstance(val, str):
                val_stripped = val.strip()
                if val_stripped == "-" or val_stripped == "":
                    continue
                # Try as number
                try:
                    pct = float(val_stripped)
                except ValueError:
                    # It's a significance letter like "b" or "BC"
                    sig = val_stripped

            cells.append({
                "cmr_ref":          None,  # filled by orchestrator from OT Sheet
                "question_text":    current_question,
                "response_option":  response_option,
                "banner_group":     bc["banner_group"],
                "banner_level":     bc["level"],
                "banner_letter":    bc["letter"],
                "base_n":           bc["base_n"],
                "pct":              pct,
                "sig_higher_than":  sig,
                "source_file":      path.name,
                "session_set":      session_set,
            })

    return pd.DataFrame(cells)


def parse_all_crosstabs() -> pd.DataFrame:
    """Parse all FFX Tables workbooks → combined crosstab_cell DataFrame."""
    print("  parsing banner crosstab workbooks...")

    # Find all FFX Tables files under the reports directory
    ffx_tables = sorted(config.FFX_REPORTS.rglob("FFX Tables set *.xlsx"))
    # Also check Foodfax Tables pattern
    ffx_tables += sorted(config.FFX_REPORTS.rglob("Foodfax Tables Set *.xlsx"))

    all_dfs = []
    for path in ffx_tables:
        # Skip archive copies
        if "Archive" in str(path):
            continue

        # Extract session set from filename
        m = re.search(r"[Ss]et\s*(\d+)", path.name)
        session_set = f"Set {m.group(1)}" if m else path.stem

        print(f"    {path.name} → {session_set}")
        df = parse_crosstab_file(path, session_set)
        if not df.empty:
            all_dfs.append(df)
            print(f"      {len(df)} cells")

    if all_dfs:
        combined = pd.concat(all_dfs, ignore_index=True)
        print(f"    total: {len(combined)} crosstab cells")
        return combined
    return pd.DataFrame()
