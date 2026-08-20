"""
Phase 2 · Parse Norm Data (6,770 × 113, header row 7).

The main FoodFax database. Contains:
  - Rows 2-6: column-level statistics (count, mean, min, max, sd, se)
  - Row 7: header
  - Rows 8+: product data
  - Columns 1-27: dimensions
  - Columns 28-55: mean scores for each measure
  - Columns 57-59: calculated scores
  - Columns 60-87: T2B%/T4B% percentages
  - Columns 88-95: frequency distribution
  - Columns 96-102: purchase intent breakdown
  - Columns 103-110: additional measures and weights
  - Columns 111-113: filters/metadata

Outputs:
  - product_test DataFrame (6,770 rows)
  - measure_value DataFrame (melted long form, ~150k rows)
  - norm_column_stat DataFrame (stats from rows 2-6)
"""
from __future__ import annotations

import re
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

import config

ARCHIVE_PATH = config.FFX_EXCEL / "Foodfax Database ARCHIVE.xlsm"

# Dimension columns (not melted into measure_value)
DIM_COLS = {
    "YEAR", "Project", "Unique Reference", "Project Name", "Project Type",
    "Prodname", "Report", "Price", "Weight/Size", "Manufacturer Code",
    "MANUFACTURER", "PRODREF", "Category code", "Category",
    "2nd Category Code", "2nd Name", "3rd Category Code", "3rd Name",
    "Storage", "Sweet or Savoury", "Type", "FAXTYPE",
    "Premium Standard Value", "OWNLABEL", "DATE", "Full report",
    "Sample size",
}

# Non-measure columns (calculated scores, weights, filters)
SKIP_COLS = {
    "Product Promoter Score",
    "Unweighted Overall Score (2dcp)", "Weighted Overall Score",
    "Grocer Innovation Score",
    "New & Diff Mean Weight", "New & Diff T2B% Weight",
    "Overall Score Weight", "Relevance Score Weight",
    "Unweighted Relevance Score (2dcp)", "Weighted Relevance Score",
    "Packaging_Filter", "Weighting Type",
    "",  # empty col 113
}

# Statistics labels in rows 2-6 of Norm Data
STAT_LABELS = ["count", "mean", "min", "max", "sd", "se"]


def _parse_variant(col_name: str) -> tuple[str, str]:
    """Extract (measure_name, variant) from a column name."""
    col = col_name.strip()
    if col.startswith("T2B% "):
        return col[5:].strip(), "T2B"
    if col.startswith("T4B% "):
        return col[5:].strip(), "T4B"
    if col.startswith("B2B% "):
        return col[5:].strip(), "B2B"
    if col.startswith("Frequency% "):
        return col, "FREQ"
    if col.startswith("Full Price PI% "):
        return col, "PI_DIST"
    if col.startswith("Pre trial PI% "):
        return col, "PI_DIST"
    return col, "MEAN"


def parse_norm_data() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Returns (products_df, measures_df, stats_df).

    products_df: one row per product (6,770 rows)
    measures_df: melted measure values (product_test_id, measure_name, variant, value)
    stats_df:    column-level statistics from rows 2-6
    """
    print("  loading Norm Data from ARCHIVE workbook...")
    wb = load_workbook(str(ARCHIVE_PATH), read_only=True, data_only=True)
    ws = wb["Norm Data"]

    # Read all rows
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 8:
        raise ValueError(f"Norm Data has only {len(rows)} rows, expected 6,777+")

    # Row 7 (0-indexed row 6) is the header
    header_raw = rows[6]
    header = [_clean(c) for c in header_raw]

    # ── Stats from rows 2-6 (0-indexed 1-5) ────────────────────────────
    stats_rows = []
    for stat_idx, stat_label in enumerate(STAT_LABELS):
        row = rows[1 + stat_idx]  # rows 2-6 are indices 1-5
        for col_idx, val in enumerate(row):
            if col_idx >= len(header):
                continue
            col_name = header[col_idx]
            if col_name in DIM_COLS or col_name in SKIP_COLS or not col_name:
                continue
            if val is not None:
                try:
                    stats_rows.append({
                        "column_name": col_name,
                        "statistic": stat_label,
                        "value": float(val),
                    })
                except (ValueError, TypeError):
                    pass

    stats_df = pd.DataFrame(stats_rows)
    print(f"    {len(stats_df)} stat values from rows 2-6")

    # ── Product data from rows 8+ (0-indexed 7+) ───────────────────────
    data_rows = rows[7:]
    print(f"    {len(data_rows)} product rows")

    # Build products DataFrame
    products = []
    for row_idx, row in enumerate(data_rows):
        if not row or row[0] is None:
            continue
        d = {}
        for col_idx, val in enumerate(row):
            if col_idx < len(header):
                d[header[col_idx]] = val
        products.append(d)

    raw_df = pd.DataFrame(products)
    print(f"    {len(raw_df)} non-empty product rows")

    # Build product_test DataFrame
    products_df = pd.DataFrame({
        "unique_reference":   raw_df.get("Unique Reference"),
        "cmr_ref":            raw_df.get("Report", pd.Series(dtype="object")).apply(_str_or_none),
        "prodref":            raw_df.get("PRODREF", pd.Series(dtype="object")).apply(_str_or_none),
        "project_code":       raw_df.get("Project"),
        "project_name":       raw_df.get("Project Name"),
        "project_type":       raw_df.get("Project Type"),
        "product_name":       raw_df.get("Prodname"),
        "category_code":      raw_df.get("Category code", pd.Series(dtype="object")).apply(_str_or_none),
        "category_name":      raw_df.get("Category"),
        "manufacturer_code":  raw_df.get("Manufacturer Code", pd.Series(dtype="object")).apply(_str_or_none),
        "manufacturer_name":  raw_df.get("MANUFACTURER"),
        "own_label_or_brand": raw_df.get("OWNLABEL"),
        "test_year":          raw_df.get("YEAR", pd.Series(dtype="object")).apply(_int_or_none),
        "test_date":          raw_df.get("DATE").apply(_date_or_none),
        "session_set":        pd.Series([None] * len(raw_df)),
        "test_order":         pd.Series([None] * len(raw_df), dtype="Int64"),
        "price_gbp":          raw_df.get("Price").apply(_float_or_none),
        "pack_size":          raw_df.get("Weight/Size"),
        "storage":            raw_df.get("Storage"),
        "sweet_or_savoury":   raw_df.get("Sweet or Savoury"),
        "fax_type":           raw_df.get("FAXTYPE"),
        "tier":               raw_df.get("Premium Standard Value"),
        "sample_size":        raw_df.get("Sample size", pd.Series(dtype="object")).apply(_int_or_none),
        "attribute_scale":    pd.Series([5] * len(raw_df), dtype="Int64"),
        "source_table":       pd.Series(["norm_data"] * len(raw_df)),
        "source_file":        pd.Series(["Foodfax Database ARCHIVE.xlsm"] * len(raw_df)),
    })

    # ── Melt measures to long form ──────────────────────────────────────
    measure_cols = [c for c in header if c and c not in DIM_COLS and c not in SKIP_COLS]
    # De-duplicate column names by keeping track of indices
    measure_col_indices = []
    for idx, c in enumerate(header):
        if c and c not in DIM_COLS and c not in SKIP_COLS:
            measure_col_indices.append((idx, c))

    measures = []
    for row_idx, row in enumerate(data_rows):
        if not row or row[0] is None:
            continue
        for col_idx, col_name in measure_col_indices:
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None:
                continue
            try:
                fval = float(val)
            except (ValueError, TypeError):
                continue
            measure_name, variant = _parse_variant(col_name)
            measures.append({
                "product_row_idx": row_idx,
                "measure_name": measure_name,
                "variant": variant,
                "value": fval,
                "source_table": "norm_data",
            })

    measures_df = pd.DataFrame(measures)
    print(f"    {len(measures_df)} measure values melted")

    return products_df, measures_df, stats_df


def parse_current_norm_summary() -> pd.DataFrame:
    """Parse Current Norm Summary sheet (198 rows × 17 cols)."""
    print("  loading Current Norm Summary...")
    wb = load_workbook(str(ARCHIVE_PATH), read_only=True, data_only=True)
    ws = wb["Current Norm Summary"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = [_clean(c) for c in rows[0]]
    data = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        d = dict(zip(header, row))
        data.append({
            "norm_label":      _str_or_none(d.get(header[0])),  # "NORM Jan 2025"
            "category_code":   None,
            "category_name":   d.get("Category Name") or d.get("Short Category Name"),
            "total":           _int_or_none(d.get("Total")),
            "premium":         _int_or_none(d.get("Premium")),
            "standard_value":  _int_or_none(d.get("Standard & Value")),
            "total_last_5y":   _int_or_none(d.get("Total (last 5 years)")),
            "pct_last_5y":     _float_or_none(d.get("Percentage (last 5 years)")),
            "y2024":           _int_or_none(d.get("2024")),
            "y2023":           _int_or_none(d.get("2023")),
            "y2022":           _int_or_none(d.get("2022")),
            "y2021":           _int_or_none(d.get("2021")),
            "y2020":           _int_or_none(d.get("2020")),
            "y2019":           _int_or_none(d.get("2019")),
            "y2018":           _int_or_none(d.get("2018")),
            "y2017":           _int_or_none(d.get("2017")),
            "y2016":           _int_or_none(d.get("2016")),
        })

    df = pd.DataFrame(data)
    print(f"    {len(df)} category norms loaded")
    return df


# ── helpers ─────────────────────────────────────────────────────────────

def _clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _str_or_none(v):
    if v is None:
        return None
    return str(v).strip() or None


def _int_or_none(v):
    if v is None:
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _float_or_none(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _date_or_none(v):
    """Convert date values to ISO string, handling bad values like 'N/A'."""
    if v is None:
        return None
    import datetime
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v).strip()
    if s.lower() in ("n/a", "na", "", "date", "none"):
        return None
    return s
